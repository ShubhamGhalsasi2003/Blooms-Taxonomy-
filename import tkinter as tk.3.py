import os
import tkinter as tk
from tkinter import messagebox
import openpyxl

class BloomGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Bloom's Taxonomy GUI")
        self.master.geometry("600x600")

        self.label = tk.Label(master, text="Enter questions (separated by commas):", bg='black', fg='white')
        self.label.pack(pady=10)

        self.question_entry = tk.Entry(master, width=70, font=("Arial", 16))
        self.question_entry.pack(pady=10)

        self.show_description_button = tk.Button(master, text="Show Descriptions", command=self.show_descriptions)
        self.show_description_button.pack(pady=20)

        self.description_label = tk.Label(master, text="", font=("Arial", 17), bg='black', fg='white')
        self.description_label.pack()

        self.questions = []
        self.levels_keywords = {
            "Remember": ["identify", "recall", "recognize", "define", "list", "memorize", "state,Who"],
            "Understand": ["explain", "compare", "interpret", "illustrate", "discuss", "predict"],
            "Apply": ["apply", "execute", "translate", "solve", "demonstrate", "utilize"],
            "Analyze": ["analyze", "compare", "categorize", "dissect"],
            "Evaluate": ["evaluate", "critique", "assess", "judge", "prioritize", "verify"],
            "Create": ["create", "design", "invent", "compose", "imagine", "generate"]
        }

    def identify_level(self, question):
        question_lower = question.lower()

        for level, keywords in self.levels_keywords.items():
            if any(keyword in question_lower for keyword in keywords):
                return level

        return "Unknown"

    def show_descriptions(self):
        questions_input = self.question_entry.get().strip()
        if not questions_input:
            messagebox.showerror("Error", "Please enter at least one question.")
            return

        self.questions = [q.strip() for q in questions_input.split(",")]

        descriptions = ""
        for idx, question in enumerate(self.questions, start=1):
            identified_level = self.identify_level(question)
            if identified_level == "Unknown":
                descriptions += f"Question {idx}: Unable to identify the level of the question\n"
            else:
                descriptions += f"Question {idx}: Category: {identified_level}\n"

        self.description_label.config(text=descriptions)

        # Create or update the Excel file
        self.update_excel()

    def update_excel(self):
        # Get the desktop directory
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

        # Form the complete file path
        file_path = os.path.join(desktop_path, "questions.xlsx")
        
        print("File path:", file_path)  # Print the file path for debugging

        # Open or create the Excel file
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Questions"

        # Write headers if the file is newly created
        if sheet.cell(row=1, column=1).value is None:
            sheet['A1'] = "Question Number"
            sheet['B1'] = "Question"
            sheet['C1'] = "Level"

        # Write questions and levels
        start_row = sheet.max_row + 1
        for idx, question in enumerate(self.questions, start=start_row):
            sheet.cell(row=idx, column=1, value=idx)
            sheet.cell(row=idx, column=2, value=question)
            level = self.identify_level(question)
            sheet.cell(row=idx, column=3, value=level)

        # Save the Excel file
        workbook.save(file_path)


class LoginInterface:
    def __init__(self, master):
        self.master = master
        self.master.title("Login Page")
        self.master.geometry("400x200")
        self.master.configure(bg="black")

        self.label = tk.Label(master, text="Login as:", bg='black', fg='white')
        self.label.pack(pady=10)

        self.student_button = tk.Button(master, text="Student", command=self.show_student_login)
        self.student_button.pack(pady=5)

        self.teacher_button = tk.Button(master, text="Teacher", command=self.show_teacher_login)
        self.teacher_button.pack(pady=5)

    def show_student_login(self):
        self.master.destroy()
        student_login_window = tk.Tk()
        student_login_page = StudentLoginPage(student_login_window)
        student_login_window.mainloop()

    def show_teacher_login(self):
        self.master.destroy()
        teacher_login_window = tk.Tk()
        teacher_login_page = TeacherLoginPage(teacher_login_window)
        teacher_login_window.mainloop()


class StudentLoginPage:
    def __init__(self, master):
        self.master = master
        self.master.title("Student Login")
        self.master.geometry("400x200")
        self.master.configure(bg='black')

        self.label = tk.Label(master, text="Student Login", bg='black', fg='white')
        self.label.pack(pady=10)

        self.username_label = tk.Label(master, text="Username:", bg='black', fg='white')
        self.username_label.pack(pady=5)

        self.username_entry = tk.Entry(master, width=30)
        self.username_entry.pack(pady=5)

        self.password_label = tk.Label(master, text="Password:", bg='black', fg='white')
        self.password_label.pack(pady=5)

        self.password_entry = tk.Entry(master, show="*", width=30)
        self.password_entry.pack(pady=5)

        self.login_button = tk.Button(master, text="Login", command=self.student_login, bg='black', fg='white')
        self.login_button.pack(pady=10)

    def student_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        if username == "student" and password == "password":
            messagebox.showinfo("Success", "Student Login Successful")
            self.show_blooms_gui()
        else:
            messagebox.showerror("Error", "Invalid Credentials")

    def show_blooms_gui(self):
        self.master.destroy()
        blooms_window = tk.Tk()
        blooms_gui = BloomGUI(blooms_window)
        blooms_window.mainloop()


class TeacherLoginPage:
    def __init__(self, master):
        self.master = master
        self.master.title("Teacher Login")
        self.master.geometry("400x200")

        self.label = tk.Label(master, text="Teacher Login")
        self.label.pack(pady=10)

        self.username_label = tk.Label(master, text="Username:")
        self.username_label.pack(pady=5)

        self.username_entry = tk.Entry(master, width=30)
        self.username_entry.pack(pady=5)

        self.password_label = tk.Label(master, text="Password:")
        self.password_label.pack(pady=5)

        self.password_entry = tk.Entry(master, show="*", width=30)
        self.password_entry.pack(pady=5)

        self.login_button = tk.Button(master, text="Login", command=self.teacher_login)
        self.login_button.pack(pady=10)

    def teacher_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        if username == "teacher" and password == "password":
            messagebox.showinfo("Success", "Teacher Login Successful")
            self.show_blooms_gui()
        else:
            messagebox.showerror("Error", "Invalid Credentials")

    def show_blooms_gui(self):
        self.master.destroy()
        blooms_window = tk.Tk()
        blooms_gui = BloomGUI(blooms_window)
        blooms_window.mainloop()


if __name__ == "__main__":
    root = tk.Tk()
    login_page = LoginInterface(root)
    root.mainloop()
